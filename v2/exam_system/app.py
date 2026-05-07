"""
在线考试系统 v2 - Online Exam System
功能：批量导入用户、程序题沙箱评测、题库管理、成绩管理、账号管理
"""

import os, sys, csv, io, json, random, sqlite3, hashlib, secrets
import textwrap, threading, webbrowser, subprocess, tempfile
from datetime import datetime, timedelta
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, g, Response)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BASE_DIR = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
DB_PATH  = os.path.join(os.path.dirname(os.path.abspath(
    sys.executable if getattr(sys, 'frozen', False) else __file__
)), 'exam_system.db')

app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'templates'),
            static_folder=os.path.join(BASE_DIR, 'static'))
app.secret_key = secrets.token_hex(32)
app.permanent_session_lifetime = timedelta(hours=2)
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024

# ═══════════════════════════════════════════════════════════
#  DB
# ═══════════════════════════════════════════════════════════
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()

def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

def init_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        username   TEXT UNIQUE NOT NULL,
        password   TEXT NOT NULL,
        role       TEXT NOT NULL DEFAULT 'student',
        name       TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS exams (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        title       TEXT NOT NULL,
        description TEXT,
        duration    INTEGER DEFAULT 60,
        pass_score  INTEGER DEFAULT 60,
        shuffle     INTEGER DEFAULT 1,
        status      TEXT DEFAULT 'draft',
        created_by  INTEGER,
        created_at  TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY(created_by) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS questions (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        exam_id        INTEGER NOT NULL,
        qtype          TEXT NOT NULL,
        content        TEXT NOT NULL,
        options        TEXT,
        answer         TEXT NOT NULL DEFAULT '',
        explanation    TEXT,
        score          INTEGER DEFAULT 10,
        sort_order     INTEGER DEFAULT 0,
        func_name      TEXT,
        func_signature TEXT,
        test_cases     TEXT,
        num_random     INTEGER DEFAULT 5,
        time_limit     INTEGER DEFAULT 5,
        FOREIGN KEY(exam_id) REFERENCES exams(id)
    );
    CREATE TABLE IF NOT EXISTS exam_records (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        exam_id     INTEGER NOT NULL,
        user_id     INTEGER NOT NULL,
        answers     TEXT,
        score       INTEGER DEFAULT 0,
        total       INTEGER DEFAULT 0,
        passed      INTEGER DEFAULT 0,
        started_at  TEXT DEFAULT (datetime('now','localtime')),
        finished_at TEXT,
        time_used   INTEGER DEFAULT 0,
        FOREIGN KEY(exam_id) REFERENCES exams(id),
        FOREIGN KEY(user_id) REFERENCES users(id)
    );
    """)
    # 迁移旧库
    for col, defn in [('func_name','TEXT'),('func_signature','TEXT'),
                      ('test_cases','TEXT'),('num_random','INTEGER DEFAULT 5'),
                      ('time_limit','INTEGER DEFAULT 5')]:
        try: db.execute(f"ALTER TABLE questions ADD COLUMN {col} {defn}")
        except: pass

    db.execute("INSERT OR IGNORE INTO users(username,password,role,name) VALUES(?,?,?,?)",
               ('admin', hash_pw('admin123'), 'admin', '系统管理员'))
    db.execute("INSERT OR IGNORE INTO users(username,password,role,name) VALUES(?,?,?,?)",
               ('student', hash_pw('student123'), 'student', '测试学生'))
    db.commit()
    _seed_demo(db)
    db.close()

def _seed_demo(db):
    if db.execute("SELECT COUNT(*) FROM exams").fetchone()[0] > 0:
        return
    aid = db.execute("SELECT id FROM users WHERE username='admin'").fetchone()[0]
    db.execute("INSERT INTO exams(title,description,duration,pass_score,shuffle,status,created_by) VALUES(?,?,?,?,?,?,?)",
               ('Python 综合测验', '涵盖选择、填空与编程题', 45, 60, 1, 'published', aid))
    eid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    rows = [
        ('single','Python 定义函数的关键字是？',
         json.dumps(['class','def','func','define']),'B','def 是函数定义关键字',10,None,None,None,5,5),
        ('single','哪个是不可变类型？',
         json.dumps(['list','dict','set','tuple']),'D','tuple 不可变',10,None,None,None,5,5),
        ('multi','哪些是 Python 内置函数？（多选）',
         json.dumps(['print()','len()','push()','range()']),'A,B,D','push() 不存在',10,None,None,None,5,5),
        ('truefalse','Python 是编译型语言。',
         json.dumps(['正确','错误']),'B','Python 是解释型',10,None,None,None,5,5),
        ('fillblank','获取列表长度的函数是 ___。',
         None,'len','len() 函数',10,None,None,None,5,5),
        ('coding',
         '实现 `solution(nums, target)`，在整数列表中找两数之和等于 target，返回下标列表（升序）。\n\n**示例：** nums=[2,7,11,15], target=9 → [0,1]',
         None,'','哈希表 O(n) 解法',30,
         'solution','def solution(nums: list, target: int) -> list:',
         json.dumps([
             {'input':[[2,7,11,15],9],'output':[0,1]},
             {'input':[[3,2,4],6],   'output':[1,2]},
             {'input':[[3,3],6],     'output':[0,1]},
         ]),5,5),
    ]
    for i,r in enumerate(rows):
        db.execute("""INSERT INTO questions(exam_id,qtype,content,options,answer,explanation,score,
                      sort_order,func_name,func_signature,test_cases,num_random,time_limit)
                      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                   (eid,)+r[:7]+(i,)+r[7:])
    db.commit()

# ═══════════════════════════════════════════════════════════
#  Auth
# ═══════════════════════════════════════════════════════════
def login_required(f):
    @wraps(f)
    def d(*a,**kw):
        if 'user_id' not in session:
            flash('请先登录','warning'); return redirect(url_for('login'))
        return f(*a,**kw)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a,**kw):
        if session.get('role') != 'admin':
            flash('权限不足','danger'); return redirect(url_for('index'))
        return f(*a,**kw)
    return login_required(d)

# ═══════════════════════════════════════════════════════════
#  代码沙箱
# ═══════════════════════════════════════════════════════════
def _gen_random_cases(fixed, n):
    """根据固定用例输入类型自动生成 n 个随机用例"""
    if not fixed: return []
    sample = fixed[0].get('input', [])
    cases = []
    for _ in range(n):
        args = []
        for a in sample:
            if isinstance(a, list):
                ln = random.randint(2, 8)
                args.append([random.randint(-20, 50) for _ in range(ln)])
            elif isinstance(a, int):
                args.append(random.randint(1, 30))
            elif isinstance(a, float):
                args.append(round(random.uniform(0, 10), 2))
            elif isinstance(a, str):
                args.append(random.choice(['hello','world','python','test']))
            else:
                args.append(a)
        cases.append({'input': args, 'output': None, 'is_random': True})
    return cases

def run_sandbox(user_code, func_name, test_cases, num_random=5, time_limit=5):
    fixed  = [c for c in test_cases if not c.get('is_random')]
    randoms = _gen_random_cases(fixed, num_random)
    all_cases = fixed + randoms

    cases_json = json.dumps(all_cases)
    fname_json = json.dumps(func_name)
    script = textwrap.dedent(f"""
import sys, json, traceback
{user_code}
cases     = json.loads({repr(cases_json)})
func_name = json.loads({repr(fname_json)})
results   = []
try:
    fn = globals()[func_name]
except KeyError:
    print(json.dumps({{"error": f"未找到函数 {{func_name}}"}})); sys.exit()
for i, c in enumerate(cases):
    args     = c["input"]
    expected = c.get("output")
    is_rand  = c.get("is_random", False)
    try:
        actual = fn(*args)
        if is_rand:
            results.append({{"idx":i,"ok":True,"is_random":True,
                             "input":args,"actual":str(actual)}})
        else:
            ok = (actual == expected)
            results.append({{"idx":i,"ok":ok,"is_random":False,
                             "input":args,"expected":expected,"actual":str(actual)}})
    except Exception:
        results.append({{"idx":i,"ok":False,"is_random":is_rand,
                         "input":args,"expected":expected,
                         "error":traceback.format_exc(limit=2)}})
print(json.dumps({{"results":results}}))
""")
    try:
        with tempfile.NamedTemporaryFile('w', suffix='.py', delete=False, encoding='utf-8') as f:
            f.write(script); tmp = f.name
        proc = subprocess.run([sys.executable, tmp], capture_output=True, text=True,
                              timeout=time_limit * len(all_cases) + 3)
        os.unlink(tmp)
        out = proc.stdout.strip()
        if not out:
            return {'passed':0,'total':len(all_cases),'details':[],
                    'error': proc.stderr.strip() or '无输出，可能有语法错误'}
        data = json.loads(out)
        if 'error' in data:
            return {'passed':0,'total':len(all_cases),'details':[],'error':data['error']}
        res = data['results']
        return {'passed':sum(1 for r in res if r.get('ok')),
                'total':len(res),'details':res,'error':proc.stderr.strip() or None}
    except subprocess.TimeoutExpired:
        try: os.unlink(tmp)
        except: pass
        return {'passed':0,'total':len(all_cases),'details':[],'error':'执行超时'}
    except Exception as e:
        return {'passed':0,'total':len(all_cases),'details':[],'error':str(e)}

# ═══════════════════════════════════════════════════════════
#  批量导入用户解析
# ═══════════════════════════════════════════════════════════
_ALIAS = {
    'username': ['username','用户名','账号','loginname','login'],
    'name':     ['name','姓名','真实姓名','realname','full_name'],
    'password': ['password','密码','passwd','pwd','pass'],
}

def _norm_header(h): return str(h).strip().lower().replace(' ','')

def _get_col(norm_row, field):
    for a in _ALIAS[field]:
        if a in norm_row: return norm_row[a]
    return ''

def parse_csv(content):
    users, errors = [], []
    reader = csv.DictReader(io.StringIO(content))
    for i, row in enumerate(reader, 2):
        nr = {_norm_header(k): str(v).strip() for k,v in row.items()}
        u = _get_col(nr,'username')
        if not u: errors.append(f'第{i}行：用户名为空，跳过'); continue
        users.append((u, _get_col(nr,'name') or u, _get_col(nr,'password') or '123456'))
    return users, errors

def parse_xlsx(data):
    users, errors = [], []
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return [], ['文件为空']
    hdrs = [_norm_header(h) if h else '' for h in rows[0]]
    def ci(field):
        for a in _ALIAS[field]:
            if a in hdrs: return hdrs.index(a)
        return -1
    iu,iname,ipw = ci('username'),ci('name'),ci('password')
    if iu < 0: return [], ['未找到"用户名"列']
    for i, row in enumerate(rows[1:], 2):
        def cell(idx): return str(row[idx]).strip() if idx>=0 and idx<len(row) and row[idx] else ''
        u = cell(iu)
        if not u: errors.append(f'第{i}行：用户名为空，跳过'); continue
        users.append((u, cell(iname) or u, cell(ipw) or '123456'))
    return users, errors

# ═══════════════════════════════════════════════════════════
#  Routes - 公共
# ═══════════════════════════════════════════════════════════
@app.route('/')
def index():
    if 'user_id' not in session: return redirect(url_for('login'))
    return redirect(url_for('admin_dashboard') if session['role']=='admin'
                    else url_for('student_dashboard'))

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        u,p = request.form['username'].strip(), request.form['password']
        row = get_db().execute(
            "SELECT * FROM users WHERE username=? AND password=?", (u, hash_pw(p))
        ).fetchone()
        if row:
            session.permanent = True
            session.update({'user_id':row['id'],'username':row['username'],
                            'name':row['name'] or row['username'],'role':row['role']})
            return redirect(url_for('index'))
        flash('用户名或密码错误','danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear(); return redirect(url_for('login'))

@app.route('/register', methods=['GET','POST'])
def register():
    if request.method == 'POST':
        u,p,n = request.form['username'].strip(),request.form['password'],request.form['name'].strip()
        if not u or not p:
            flash('用户名密码不能为空','danger')
        else:
            try:
                get_db().execute("INSERT INTO users(username,password,role,name) VALUES(?,?,?,?)",
                                 (u, hash_pw(p), 'student', n))
                get_db().commit(); flash('注册成功，请登录','success')
                return redirect(url_for('login'))
            except sqlite3.IntegrityError:
                flash('用户名已存在','danger')
    return render_template('register.html')

# ═══════════════════════════════════════════════════════════
#  Routes - 学生端
# ═══════════════════════════════════════════════════════════
@app.route('/student')
@login_required
def student_dashboard():
    db = get_db()
    exams = db.execute("SELECT * FROM exams WHERE status='published' ORDER BY created_at DESC").fetchall()
    recs  = {r['exam_id']:r for r in db.execute(
        "SELECT * FROM exam_records WHERE user_id=? ORDER BY started_at DESC",(session['user_id'],)).fetchall()}
    return render_template('student_dashboard.html', exams=exams, records=recs)

@app.route('/exam/<int:exam_id>/start')
@login_required
def start_exam(exam_id):
    db   = get_db()
    exam = db.execute("SELECT * FROM exams WHERE id=? AND status='published'",(exam_id,)).fetchone()
    if not exam: flash('考试不存在或未开放','warning'); return redirect(url_for('student_dashboard'))
    rec = db.execute("SELECT * FROM exam_records WHERE exam_id=? AND user_id=? AND finished_at IS NULL",
                     (exam_id, session['user_id'])).fetchone()
    if not rec:
        db.execute("INSERT INTO exam_records(exam_id,user_id) VALUES(?,?)",(exam_id,session['user_id']))
        db.commit()
        rec = db.execute("SELECT * FROM exam_records WHERE exam_id=? AND user_id=? AND finished_at IS NULL",
                         (exam_id,session['user_id'])).fetchone()
    qs = [dict(q) for q in db.execute(
        "SELECT * FROM questions WHERE exam_id=? ORDER BY sort_order",(exam_id,)).fetchall()]
    for q in qs:
        if q['options']:    q['options']    = json.loads(q['options'])
        if q.get('test_cases'): q['test_cases'] = json.loads(q['test_cases'])
    if exam['shuffle']:
        nc = [q for q in qs if q['qtype']!='coding']
        cq = [q for q in qs if q['qtype']=='coding']
        random.shuffle(nc); qs = nc + cq
    return render_template('exam_take.html', exam=exam, questions=qs, record=rec)

@app.route('/exam/<int:rid>/submit', methods=['POST'])
@login_required
def submit_exam(rid):
    db  = get_db()
    rec = db.execute("SELECT * FROM exam_records WHERE id=? AND user_id=? AND finished_at IS NULL",
                     (rid, session['user_id'])).fetchone()
    if not rec: flash('提交失败','danger'); return redirect(url_for('student_dashboard'))

    qs    = db.execute("SELECT * FROM questions WHERE exam_id=?",(rec['exam_id'],)).fetchall()
    total = sum(q['score'] for q in qs)
    score = 0; detail = {}

    for q in qs:
        qid = str(q['id'])
        if q['qtype'] == 'coding':
            code = request.form.get(f'code_{q["id"]}','').strip()
            if not code:
                detail[qid]={'type':'coding','user':'','ok':False,'score':q['score'],'earned':0,'sandbox':None}
                continue
            tc  = json.loads(q['test_cases']) if q['test_cases'] else []
            res = run_sandbox(code, q['func_name'], tc, q['num_random'] or 5, q['time_limit'] or 5)
            # 只按固定用例计分，随机用例只作为运行稳定性参考
            fixed_details = [d for d in res['details'] if not d.get('is_random')]
            fixed_passed  = sum(1 for d in fixed_details if d.get('ok'))
            fixed_total   = len(fixed_details) if fixed_details else 1
            ratio  = fixed_passed / fixed_total
            earned = round(q['score'] * ratio)
            res['fixed_passed'] = fixed_passed
            res['fixed_total']  = fixed_total
            score += earned
            detail[qid]={'type':'coding','user':code,'ok':ratio>=1.0,'score':q['score'],'earned':earned,'sandbox':res}
        elif q['qtype'] == 'multi':
            ua = ','.join(sorted(request.form.getlist(f'm_{q["id"]}')))
            ok = ua == q['answer'].strip().upper()
            if ok: score += q['score']
            detail[qid]={'user':ua,'correct':q['answer'],'ok':ok,'score':q['score'],'type':q['qtype']}
        elif q['qtype'] == 'fillblank':
            ua = request.form.get(f'q_{q["id"]}','').strip()
            ok = ua.lower() == q['answer'].strip().lower()
            if ok: score += q['score']
            detail[qid]={'user':ua,'correct':q['answer'],'ok':ok,'score':q['score'],'type':q['qtype']}
        else:
            ua = request.form.get(f'q_{q["id"]}','').strip().upper()
            ok = ua == q['answer'].strip().upper()
            if ok: score += q['score']
            detail[qid]={'user':ua,'correct':q['answer'],'ok':ok,'score':q['score'],'type':q['qtype']}

    exam   = db.execute("SELECT * FROM exams WHERE id=?",(rec['exam_id'],)).fetchone()
    passed = 1 if total and (score/total*100) >= exam['pass_score'] else 0
    used   = int((datetime.now()-datetime.strptime(rec['started_at'],'%Y-%m-%d %H:%M:%S')).total_seconds())
    db.execute("UPDATE exam_records SET answers=?,score=?,total=?,passed=?,"
               "finished_at=datetime('now','localtime'),time_used=? WHERE id=?",
               (json.dumps(detail),score,total,passed,used,rid))
    db.commit()
    return redirect(url_for('exam_result', record_id=rid))

@app.route('/result/<int:record_id>')
@login_required
def exam_result(record_id):
    db  = get_db()
    rec = db.execute("SELECT er.*,e.title,e.pass_score,e.duration FROM exam_records er "
                     "JOIN exams e ON er.exam_id=e.id WHERE er.id=? AND er.user_id=?",
                     (record_id,session['user_id'])).fetchone()
    if not rec: flash('记录不存在','warning'); return redirect(url_for('student_dashboard'))
    qs = [dict(q) for q in db.execute(
        "SELECT * FROM questions WHERE exam_id=? ORDER BY sort_order",(rec['exam_id'],)).fetchall()]
    for q in qs:
        if q['options']: q['options']=json.loads(q['options'])
    return render_template('exam_result.html', rec=rec, questions=qs,
                           answers=json.loads(rec['answers'] or '{}'))

@app.route('/api/run_code', methods=['POST'])
@login_required
def api_run_code():
    d = request.get_json()
    if not d.get('code','').strip(): return jsonify({'error':'代码为空'})
    return jsonify(run_sandbox(d['code'], d.get('func_name','solution'),
                               d.get('test_cases',[]), int(d.get('num_random',3)),
                               int(d.get('time_limit',5))))

# ═══════════════════════════════════════════════════════════
#  Routes - 管理端
# ═══════════════════════════════════════════════════════════
@app.route('/admin')
@admin_required
def admin_dashboard():
    db = get_db()
    tr = db.execute("SELECT COUNT(*) FROM exam_records WHERE finished_at IS NOT NULL").fetchone()[0]
    pr = db.execute("SELECT COUNT(*) FROM exam_records WHERE passed=1").fetchone()[0]
    stats = {
        'exams':    db.execute("SELECT COUNT(*) FROM exams").fetchone()[0],
        'users':    db.execute("SELECT COUNT(*) FROM users WHERE role='student'").fetchone()[0],
        'records':  tr,
        'pass_rate':round(pr/tr*100,1) if tr else 0,
        'questions':db.execute("SELECT COUNT(*) FROM questions").fetchone()[0],
    }
    recent = db.execute(
        "SELECT er.*,u.name uname,e.title etitle FROM exam_records er "
        "JOIN users u ON er.user_id=u.id JOIN exams e ON er.exam_id=e.id "
        "WHERE er.finished_at IS NOT NULL ORDER BY er.finished_at DESC LIMIT 8"
    ).fetchall()
    return render_template('admin_dashboard.html', stats=stats, recent=recent)

# ── 考试管理 ─────────────────────────────────────────────
@app.route('/admin/exams')
@admin_required
def admin_exams():
    exams = get_db().execute(
        "SELECT e.*,COUNT(q.id) qcount,u.name creator FROM exams e "
        "LEFT JOIN questions q ON e.id=q.exam_id "
        "LEFT JOIN users u ON e.created_by=u.id "
        "GROUP BY e.id ORDER BY e.created_at DESC"
    ).fetchall()
    return render_template('admin_exams.html', exams=exams)

@app.route('/admin/exam/new', methods=['GET','POST'])
@app.route('/admin/exam/<int:exam_id>/edit', methods=['GET','POST'])
@admin_required
def admin_exam_edit(exam_id=None):
    db   = get_db()
    exam = db.execute("SELECT * FROM exams WHERE id=?",(exam_id,)).fetchone() if exam_id else None
    if request.method == 'POST':
        t,desc,dur,ps,sh = (request.form['title'].strip(),
                            request.form['description'].strip(),
                            int(request.form.get('duration',60)),
                            int(request.form.get('pass_score',60)),
                            1 if request.form.get('shuffle') else 0)
        if exam_id:
            db.execute("UPDATE exams SET title=?,description=?,duration=?,pass_score=?,shuffle=? WHERE id=?",
                       (t,desc,dur,ps,sh,exam_id))
        else:
            db.execute("INSERT INTO exams(title,description,duration,pass_score,shuffle,created_by) VALUES(?,?,?,?,?,?)",
                       (t,desc,dur,ps,sh,session['user_id']))
            exam_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.commit(); flash('保存成功','success')
        return redirect(url_for('admin_exam_questions', exam_id=exam_id))
    return render_template('admin_exam_edit.html', exam=exam)

@app.route('/admin/exam/<int:exam_id>/status/<action>', methods=['POST'])
@admin_required
def admin_exam_status(exam_id, action):
    s = {'publish':'published','close':'closed','draft':'draft'}.get(action)
    if s:
        get_db().execute("UPDATE exams SET status=? WHERE id=?",(s,exam_id))
        get_db().commit(); flash('状态已更新','success')
    return redirect(url_for('admin_exams'))

@app.route('/admin/exam/<int:exam_id>/delete', methods=['POST'])
@admin_required
def admin_exam_delete(exam_id):
    db = get_db()
    db.execute("DELETE FROM questions WHERE exam_id=?",(exam_id,))
    db.execute("DELETE FROM exam_records WHERE exam_id=?",(exam_id,))
    db.execute("DELETE FROM exams WHERE id=?",(exam_id,))
    db.commit(); flash('考试已删除','success')
    return redirect(url_for('admin_exams'))

# ── 题库管理（全局） ────────────────────────────────────────
@app.route('/admin/questions')
@admin_required
def admin_question_bank():
    db = get_db()
    exam_id  = request.args.get('exam_id', type=int)
    qtype    = request.args.get('qtype','')
    keyword  = request.args.get('keyword','').strip()
    q = ("SELECT q.*,e.title exam_title FROM questions q "
         "JOIN exams e ON q.exam_id=e.id WHERE 1=1 ")
    params = []
    if exam_id: q+="AND q.exam_id=? "; params.append(exam_id)
    if qtype:   q+="AND q.qtype=? ";   params.append(qtype)
    if keyword: q+="AND q.content LIKE ? "; params.append(f'%{keyword}%')
    q += "ORDER BY e.id,q.sort_order"
    qs    = [dict(r) for r in db.execute(q, params).fetchall()]
    for r in qs:
        if r['options']: r['options']=json.loads(r['options'])
        if r.get('test_cases'): r['test_cases']=json.loads(r['test_cases'])
    exams = db.execute("SELECT id,title FROM exams ORDER BY title").fetchall()
    return render_template('admin_question_bank.html', questions=qs, exams=exams,
                           sel_exam=exam_id, sel_type=qtype, keyword=keyword)

@app.route('/admin/exam/<int:exam_id>/questions')
@admin_required
def admin_exam_questions(exam_id):
    db   = get_db()
    exam = db.execute("SELECT * FROM exams WHERE id=?",(exam_id,)).fetchone()
    qs   = [dict(q) for q in db.execute(
        "SELECT * FROM questions WHERE exam_id=? ORDER BY sort_order",(exam_id,)).fetchall()]
    for q in qs:
        if q['options']:        q['options']   = json.loads(q['options'])
        if q.get('test_cases'): q['test_cases'] = json.loads(q['test_cases'])
    return render_template('admin_exam_questions.html', exam=exam, questions=qs,
                           total_score=sum(q['score'] for q in qs))

@app.route('/admin/exam/<int:exam_id>/question/new', methods=['GET','POST'])
@app.route('/admin/exam/<int:exam_id>/question/<int:qid>/edit', methods=['GET','POST'])
@admin_required
def admin_question_edit(exam_id, qid=None):
    db   = get_db()
    exam = db.execute("SELECT * FROM exams WHERE id=?",(exam_id,)).fetchone()
    q    = None
    if qid:
        row = db.execute("SELECT * FROM questions WHERE id=? AND exam_id=?",(qid,exam_id)).fetchone()
        if row:
            q = dict(row)
            if q['options']:        q['options']   = json.loads(q['options'])
            if q.get('test_cases'): q['test_cases'] = json.loads(q['test_cases'])
    if request.method == 'POST':
        qtype   = request.form['qtype']
        content = request.form['content'].strip()
        exp     = request.form.get('explanation','').strip()
        sc      = int(request.form.get('score',10))
        back_to = request.form.get('back_to','exam')  # 'exam' or 'bank'
        if qtype == 'coding':
            fn  = request.form.get('func_name','solution').strip()
            sig = request.form.get('func_signature','').strip()
            nr  = int(request.form.get('num_random',5))
            tl  = int(request.form.get('time_limit',5))
            try:   tc = json.loads(request.form.get('test_cases_json','[]').strip())
            except: flash('测试用例 JSON 格式错误','danger'); return redirect(request.url)
            fields = (qtype,content,None,'',exp,sc,fn,sig,json.dumps(tc),nr,tl)
            if qid:
                db.execute("UPDATE questions SET qtype=?,content=?,options=?,answer=?,explanation=?,"
                           "score=?,func_name=?,func_signature=?,test_cases=?,num_random=?,time_limit=? WHERE id=?",
                           fields+(qid,))
            else:
                mo = db.execute("SELECT MAX(sort_order) FROM questions WHERE exam_id=?",(exam_id,)).fetchone()[0] or 0
                db.execute("INSERT INTO questions(qtype,content,options,answer,explanation,score,"
                           "func_name,func_signature,test_cases,num_random,time_limit,exam_id,sort_order) "
                           "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)", fields+(exam_id,mo+1))
        else:
            answer = request.form['answer'].strip()
            opts   = None
            if qtype in ('single','multi','truefalse'):
                raw  = [request.form.get(f'opt_{c}','').strip() for c in 'ABCD']
                opts = json.dumps([o for o in raw if o])
            if qid:
                db.execute("UPDATE questions SET qtype=?,content=?,options=?,answer=?,explanation=?,"
                           "score=?,func_name=NULL,func_signature=NULL,test_cases=NULL WHERE id=?",
                           (qtype,content,opts,answer,exp,sc,qid))
            else:
                mo = db.execute("SELECT MAX(sort_order) FROM questions WHERE exam_id=?",(exam_id,)).fetchone()[0] or 0
                db.execute("INSERT INTO questions(qtype,content,options,answer,explanation,score,exam_id,sort_order) "
                           "VALUES(?,?,?,?,?,?,?,?)",(qtype,content,opts,answer,exp,sc,exam_id,mo+1))
        db.commit(); flash('题目已保存','success')
        if back_to == 'bank':
            return redirect(url_for('admin_question_bank'))
        return redirect(url_for('admin_exam_questions', exam_id=exam_id))
    from_bank = request.args.get('from_bank','0')
    return render_template('admin_question_edit.html', exam=exam, q=q, from_bank=from_bank)

@app.route('/admin/question/<int:qid>/delete', methods=['POST'])
@admin_required
def admin_question_delete(qid):
    db = get_db()
    row = db.execute("SELECT exam_id FROM questions WHERE id=?",(qid,)).fetchone()
    if row:
        db.execute("DELETE FROM questions WHERE id=?",(qid,))
        db.commit(); flash('题目已删除','success')
        back = request.form.get('back_to','exam')
        if back == 'bank': return redirect(url_for('admin_question_bank'))
        return redirect(url_for('admin_exam_questions', exam_id=row['exam_id']))
    return redirect(url_for('admin_question_bank'))

# ── 成绩管理 ─────────────────────────────────────────────
@app.route('/admin/records')
@admin_required
def admin_records():
    db      = get_db()
    exam_id = request.args.get('exam_id', type=int)
    uid     = request.args.get('user_id', type=int)
    q = ("SELECT er.*,u.name uname,u.username,e.title etitle,e.pass_score "
         "FROM exam_records er JOIN users u ON er.user_id=u.id "
         "JOIN exams e ON er.exam_id=e.id WHERE er.finished_at IS NOT NULL ")
    params = []
    if exam_id: q+="AND er.exam_id=? "; params.append(exam_id)
    if uid:     q+="AND er.user_id=? "; params.append(uid)
    q += "ORDER BY er.finished_at DESC"
    records = db.execute(q, params).fetchall()
    exams   = db.execute("SELECT id,title FROM exams ORDER BY title").fetchall()
    users   = db.execute("SELECT id,name,username FROM users WHERE role='student' ORDER BY name").fetchall()
    return render_template('admin_records.html', records=records, exams=exams,
                           users=users, sel_exam=exam_id, sel_user=uid)

@app.route('/admin/record/<int:rid>/detail')
@admin_required
def admin_record_detail(rid):
    db  = get_db()
    rec = db.execute("SELECT er.*,u.name uname,e.title etitle,e.pass_score "
                     "FROM exam_records er JOIN users u ON er.user_id=u.id "
                     "JOIN exams e ON er.exam_id=e.id WHERE er.id=?",(rid,)).fetchone()
    if not rec: flash('记录不存在','warning'); return redirect(url_for('admin_records'))
    qs = [dict(q) for q in db.execute(
        "SELECT * FROM questions WHERE exam_id=? ORDER BY sort_order",(rec['exam_id'],)).fetchall()]
    for q in qs:
        if q['options']: q['options']=json.loads(q['options'])
    return render_template('admin_record_detail.html', rec=rec, questions=qs,
                           answers=json.loads(rec['answers'] or '{}'))

# ── 用户管理 + 批量导入 ──────────────────────────────────
@app.route('/admin/users')
@admin_required
def admin_users():
    keyword = request.args.get('keyword','').strip()
    role    = request.args.get('role','')
    q = ("SELECT u.*,COUNT(er.id) exam_count FROM users u "
         "LEFT JOIN exam_records er ON u.id=er.user_id WHERE 1=1 ")
    params = []
    if keyword: q+="AND (u.username LIKE ? OR u.name LIKE ?) "; params+=[f'%{keyword}%']*2
    if role:    q+="AND u.role=? "; params.append(role)
    q += "GROUP BY u.id ORDER BY u.created_at DESC"
    users = get_db().execute(q, params).fetchall()
    return render_template('admin_users.html', users=users,
                           has_openpyxl=HAS_OPENPYXL, keyword=keyword, sel_role=role)

@app.route('/admin/users/import', methods=['GET','POST'])
@admin_required
def admin_users_import():
    if request.method == 'GET':
        return render_template('admin_users_import.html', has_openpyxl=HAS_OPENPYXL)
    f = request.files.get('file')
    if not f or not f.filename:
        flash('请选择文件','danger'); return redirect(request.url)
    fname = f.filename.lower()
    try:
        if fname.endswith('.csv'):
            to_add, errs = parse_csv(f.read().decode('utf-8-sig'))
        elif fname.endswith(('.xlsx','.xls')):
            if not HAS_OPENPYXL:
                flash('未安装 openpyxl，请用 CSV','danger'); return redirect(request.url)
            to_add, errs = parse_xlsx(f.read())
        else:
            flash('只支持 .csv / .xlsx','danger'); return redirect(request.url)
    except Exception as e:
        flash(f'解析失败：{e}','danger'); return redirect(request.url)

    db = get_db(); added, skipped = [], []
    for u,n,p in to_add:
        try:
            db.execute("INSERT INTO users(username,password,role,name) VALUES(?,?,?,?)",
                       (u,hash_pw(p),'student',n)); added.append(u)
        except sqlite3.IntegrityError:
            skipped.append(u)
    db.commit()
    msg = f'✅ 成功导入 {len(added)} 个用户'
    if skipped: msg += f'  ⚠️ {len(skipped)} 个已存在跳过'
    if errs:    msg += '  |  ' + '；'.join(errs[:3])
    flash(msg, 'success' if added else 'warning')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/template/<fmt>')
@admin_required
def admin_users_template(fmt):
    if fmt == 'csv':
        c = "username,name,password\nstudent01,张三,123456\nstudent02,李四,123456\n"
        return Response(c, mimetype='text/csv',
                        headers={'Content-Disposition':'attachment;filename=import_template.csv'})
    if fmt == 'xlsx' and HAS_OPENPYXL:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title='用户列表'
        ws.append(['username','name','password'])
        ws.append(['student01','张三','123456'])
        ws.append(['student02','李四','123456'])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.read(),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition':'attachment;filename=import_template.xlsx'})
    flash('不支持该格式','danger'); return redirect(url_for('admin_users'))

@app.route('/admin/user/add', methods=['GET','POST'])
@admin_required
def admin_user_add():
    if request.method == 'POST':
        u = request.form['username'].strip()
        p = request.form['password'].strip() or '123456'
        n = request.form['name'].strip()
        r = request.form.get('role','student')
        try:
            get_db().execute("INSERT INTO users(username,password,role,name) VALUES(?,?,?,?)",
                             (u,hash_pw(p),r,n))
            get_db().commit(); flash(f'用户 {u} 创建成功','success')
            return redirect(url_for('admin_users'))
        except sqlite3.IntegrityError:
            flash('用户名已存在','danger')
    return render_template('admin_user_add.html')

@app.route('/admin/user/<int:uid>/edit', methods=['GET','POST'])
@admin_required
def admin_user_edit(uid):
    db   = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?",(uid,)).fetchone()
    if not user: flash('用户不存在','warning'); return redirect(url_for('admin_users'))
    if request.method == 'POST':
        n = request.form['name'].strip()
        r = request.form.get('role','student')
        p = request.form.get('password','').strip()
        if uid == session['user_id'] and r != 'admin':
            flash('不能修改自己的角色','danger')
        else:
            if p:
                db.execute("UPDATE users SET name=?,role=?,password=? WHERE id=?",(n,r,hash_pw(p),uid))
            else:
                db.execute("UPDATE users SET name=?,role=? WHERE id=?",(n,r,uid))
            db.commit(); flash('用户信息已更新','success')
            return redirect(url_for('admin_users'))
    return render_template('admin_user_edit.html', user=user)

@app.route('/admin/user/<int:uid>/delete', methods=['POST'])
@admin_required
def admin_user_delete(uid):
    if uid == session['user_id']:
        flash('不能删除自己','danger')
    else:
        db = get_db()
        db.execute("DELETE FROM exam_records WHERE user_id=?",(uid,))
        db.execute("DELETE FROM users WHERE id=?",(uid,))
        db.commit(); flash('用户已删除','success')
    return redirect(url_for('admin_users'))

@app.route('/admin/user/<int:uid>/reset_pw', methods=['POST'])
@admin_required
def admin_reset_pw(uid):
    pw = request.form.get('new_pw','123456').strip() or '123456'
    get_db().execute("UPDATE users SET password=? WHERE id=?",(hash_pw(pw),uid))
    get_db().commit(); flash(f'密码已重置为：{pw}','success')
    return redirect(url_for('admin_users'))

# ═══════════════════════════════════════════════════════════
if __name__ == '__main__':
    init_db()
    port = 5000
    print(f"\n{'='*52}")
    print(f"  在线考试系统 v2 启动成功")
    print(f"  地址: http://127.0.0.1:{port}")
    print(f"  管理员: admin / admin123")
    print(f"  学生:   student / student123")
    print(f"{'='*52}\n")
    threading.Timer(1.2, lambda: webbrowser.open(f'http://127.0.0.1:{port}')).start()
    app.run(host='0.0.0.0', port=port, debug=False)
